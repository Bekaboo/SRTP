# -*- coding: utf-8 -*-
"""
Created on Mon Jul 11 14:59:18 2022
Cited from: https://blog.csdn.net/ssson/article/details/118381510
"""

import openpyxl
import xlwt
import csv
 
def csv2xls(filename):      # Convert csv to xls
    with open(filename, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('data')  # Create a sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        l = 0
        for line in read:
            r = 0
            for i in line:
                sheet.write(l, r, i)  # Writing lattice by lattice
                r = r + 1
            l = l + 1
        xlsfilename = filename[:-3] + 'xls'
        workbook.save(xlsfilename)
        
def csv2xlsx(filename):     # Convert csv to xlsx
    with open(filename, 'r', encoding='utf-8') as f:
        read = csv.reader(f)
        wb = openpyxl.Workbook()
        ws = wb.active
        for line in read:
            ws.append(line)
        xlsxfilename = filename[:-3] + 'xlsx'
        wb.save(xlsxfilename)  